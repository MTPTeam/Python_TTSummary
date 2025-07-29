import openpyxl
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import time
import tkinter as tk
# from datetime import time
from datetime import date, timedelta, datetime
import os

import sys
import logging
import traceback

OpenWorkbook = UserInput = ProcessDoneMessagebox = False
ProcessDoneMessagebox = True
OpenWorkbook = True
UserInput = True




try:

    if UserInput:
        # pass
        # from tkinter import TK
        # import tkinter as tk
        root = tk.Tk()
        name_var=tk.StringVar()
        passw_var=tk.StringVar()
         
          
        root.geometry("350x450")
        root.title("Select dates")
        sdate_str = ''
        edate_str = ''
        closures = []
        e = tk.Entry(root, font=('Calibri',20))
        # e.pack()
        
        labelrow = 10
        
        def myClick():
            global labelrow
            
            inp = e.get()
            stationslice = ','.join(list(map(str.upper,[inp]))) if len(inp) > 3 else inp.upper()
            myLabel = tk.Label(root, text=f'{stationslice} added to slice')
            myLabel.grid( row=labelrow, column=0, columnspan = 2)
            closures.extend(e.get().split(','))
            labelrow+=1
            
        
        def submit():
            global sdate_str
            global edate_str
            
            closures.extend(e.get().split(','))
            sdate_str=name_var.get()
            edate_str=passw_var.get()
            root.quit()
             
             
        instr1 = tk.Label(root, text = "Closure period - Enter dates in d/m/y or d-m-y format")
        # instr2 = tk.Label(root, text = "Enter all closed station abbreviations, separated by commas")
        instr2 = tk.Label(root, text = "Affected stations - Use abbreviations separated by commas")
        blank1 = tk.Label(root, text = " ")
        # blank2 = tk.Label(root, text = " ")
        blank3 = tk.Label(root, text = " ")
        lab1  = tk.Label(root, text = 'Start Date', font=('calibre',10, 'bold'))
        lab2  = tk.Label(root, text = 'End Date', font = ('calibre',10,'bold'))
        entr1 = tk.Entry(root,textvariable = name_var, font=('calibre',10,'normal'))
        entr2 = tk.Entry(root, textvariable = passw_var, font = ('calibre',10,'normal'))
        addst = tk.Button(root,width=35, height=2, text='Add station(s)',command=myClick)
        done  = tk.Button(root,width=35,height=2,text = 'Done', command = submit)
        # done  = tk.Button(root,text = 'Done', command = submit)
        
        
        blank1.grid( row=3, column=0, columnspan = 2)
        # blank2.grid( row=7, column=0, columnspan = 2)
        blank3.grid( row=8, column=0, columnspan = 2)
        
        
        instr1.grid( row=0, column=0, columnspan = 2)
        instr2.grid( row=4, column=0, columnspan = 2)
        
        lab1.grid( row=1,column=0)
        lab2.grid( row=2,column=0)
        
        entr1.grid(row=1,column=1)
        entr2.grid(row=2,column=1)
        
        
        e.grid(    row=6,column=0, columnspan = 2)
        addst.grid(row=7,column=0, columnspan = 2)
        done.grid( row=9,column=0, columnspan = 2)
    
        root.mainloop()
        root.withdraw()
        
        date_format = '%d/%m/%y' if '/' in sdate_str else '%d-%m-%y'
        sdate = datetime.strptime(sdate_str, date_format).date()
        edate_inclusive = datetime.strptime(edate_str, date_format).date()
        
        closures = map(str.upper,closures)
        closures = map(str.strip,closures)
        closures = list(set(closures))
        
        
    else:
        sdate           = date(2024,4,2)   # start date
        edate_inclusive = date(2024,4,10)   # end date
        closures = ['DUP', 'FFI', 'YRG', 'YLY', 'MQK', 'RKE', 'SLY', 'SYK', 'ATI', 'RUC', 'FTG', 'KRY', 'TDP', 'WOI', 'KGT', 'LGL', 'BTI', 'EDL', 'HVW'] 
    
    
    
    print('Stations:  ',closures)
    print('Start Date:',sdate)
    print('End Date:  ',edate_inclusive)
    edate = edate_inclusive + timedelta(days=1)
    
    
    
    'C:\WPy64-3740\Loading Data'
    
    # myldpath = 'C:/WPy64-3740/Loading Data'
    # myldpath = 'C:/WPy64-3740'
    
    file_path = os.path.realpath(__file__)
    directory = '/'.join(file_path.split('\\')[0:-1])
    ttts = directory != 'C:/Users/r913332/OneDrive - Queensland Rail/04 Project Python/15 Closure Impacts'
    
    if ttts:
        os.chdir('//Cptprdfps001/ServicePlan/SMTP/02 PROJECTS/WPy64-3740/Closure Impacts/Loading Data')
        
    mth_loadingdata = load_workbook('PaxMoves-M2210 v9.10_Full Week (MTh).xlsx')    
    fri_loadingdata = load_workbook('PaxMoves-M2210 v9.10_Full Week (Fri).xlsx')    
    sat_loadingdata = load_workbook('PaxMoves-M2210 v9.10_Full Week (Sat).xlsx')    
    sun_loadingdata = load_workbook('PaxMoves-M2210 v9.10_Full Week (Sun).xlsx')
    mth_ws = mth_loadingdata['Data']
    fri_ws = fri_loadingdata['Data']
    sat_ws = sat_loadingdata['Data']
    sun_ws = sun_loadingdata['Data']
    
    # sys.exit()
    if ttts:
        os.chdir('//Cptprdfps001/ServicePlan/SMTP/02 PROJECTS/WPy64-3740/Closure Impacts/Impact Reports')
    
    stationmaster = {
        'Fortitude Valley':'BRC',
        'Electric Train South': 'ETS',
        'Elec Train S': 'ETS',
        'Campbell St': 'CAM',
        'Exhibition ':'EXH',
        'Exhibition': 'EXH',
        'Normanby': 'NBY',
        'Roma Street': 'RS',
        'Central': 'BNC',
        'Brunswick Street': 'BRC',
        'Bowen Hills': 'BHI',
        'Mayne': 'MNE',
        'Albion': 'AIN',
        'Wooloowin': 'WWI',
        'Eagle Junction': 'EGJ',
        'Toombul': 'TBU',
        'Nundah': 'NND',
        'Northgate': 'NTG',
        'Bindha': 'BHA',
        'Banyo': 'BQY',
        'Nudgee': 'NUD',
        'Boondall': 'BZL',
        'North Boondall': 'NBD',
        'Deagon': 'DEG',
        'Sandgate': 'SGE',
        'Shorncliffe': 'SHC',
        'Caboolture East Yard': 'CAE',
        'Caboolture': 'CAB',
        'Caboolture North': 'CEN',
        'Elimbah Stabling Yard': 'EMHS',
        'Kippa-Ring Stabling Yard': 'KPRS',
        'Kippa-Ring': 'KPR',
        'Airport Junction': 'AJN',
        'Virginia': 'VGI',
        'Sunshine': 'SSN',
        'Geebung': 'GEB',
        'Zillmere': 'ZLL',
        'Carseldine': 'CDE',
        'Bald Hills': 'BDS',
        'Strathpine': 'SPN',
        'Bray Park': 'BPR',
        'Lawnton': 'LWO',
        'Petrie': 'PET',
        'Dakabin': 'DKB',
        'Narangba': 'NRB',
        'Burpengary': 'BPY',
        'Morayfield': 'MYE',
        'Mayne North Yard': 'YN',
        'Mayne North':'YN', #!!!
        'Mayne Yard Arrival': 'YNA',
        'Roma St West Junction': 'RSWJ',
        'South Brisbane': 'SBE',
        'South Bank': 'SBA',
        'Park Road': 'PKR',
        'Dutton Park': 'DUP',
        'Fairfield': 'FFI',
        'Yeronga': 'YRG',
        'Yeerongpilly': 'YLY',
        'Moorooka': 'MQK',
        'Rocklea': 'RKE',
        'Salisbury': 'SLY',
        'Coopers Plains': 'CEP',
        'Banoon': 'BQO',
        'Sunnybank': 'SYK',
        'Altandi': 'ATI',
        'Runcorn': 'RUC',
        'Fruitgrove': 'FTG',
        'Kuraby': 'KRY',
        'Trinder Park': 'TDP',
        'Woodridge': 'WOI',
        'Kingston': 'KGT',
        'Loganlea': 'LGL',
        'Bethania': 'BTI',
        'Edens Landing': 'EDL',
        'Eden\'s Landing': 'EDL',
        'Eden’s Landing': 'EDL',
        'Holmview': 'HVW',
        'Beenleigh': 'BNH',
        'Beenleigh Turnback': 'BNT',
        'Electric Train Flyover': 'ETF',
        'Elec Train Flyover':'ETF',
        'Electric Depot Junction': 'EDJ',
        'Ipswich Stabling':'IPSS',
        'Ipswich Stabling Yard': 'IPSS',
        'Ipswich Stable':'IPSS',
        'Ipswich': 'IPS',
        'Milton': 'MTZ',
        'Auchenflower': 'AHF',
        'Toowong': 'TWG',
        'Taringa': 'TIQ',
        'Indooroopilly': 'IDP',
        'Chelmer': 'CMZ',
        'Graceville': 'GVQ',
        'Sherwood': 'SHW',
        'Corinda': 'CQD',
        'Oxley': 'OXL',
        'Darra': 'DAR',
        'Wacol': 'WAC',
        'Gailes': 'GAI',
        'Goodna': 'GDQ',
        'Redbank': 'RDK',
        'Riverview': 'RVV',
        'Dinmore': 'DIR',
        'Ebbw Vale': 'EBV',
        'Bundamba': 'BDX',
        'Booval': 'BOV',
        'East Ipswich': 'EIP',
        'Rothwell': 'RWL',
        'Mango Hill East': 'MGE',
        'Mango Hill': 'MGH',
        'Murrumba Downs': 'MRD',
        'Kallangur': 'KGR',
        'Richlands': 'RHD',
        'Springfield': 'SFD',
        'Springfield Central': 'SFC',
        'Thomas Street': 'THS',
        'Wulkuraka': 'WUL',
        'Karrabin': 'KRA',
        'Walloon': 'WOQ',
        'Thagoona': 'TAO',
        'Yarrowlea': 'YLE',
        'Rosewood': 'RSW',
        'Buranda': 'BRD',
        'Coorparoo': 'CRO',
        'Norman Park': 'NPR',
        'Morningside': 'MGS',
        'Cannon Hill': 'CNQ',
        'Murarrie': 'MJE',
        'Hemmant': 'HMM',
        'Lindum': 'LDM',
        'Lytton Junction': 'LJN',
        'Wynnum North': 'WYH',
        'Wynnum': 'WNM',
        'Wynnum Central': 'WNC',
        'Manly': 'MNY',
        'Lota': 'LOT',
        'Thorneside': 'TNS',
        'Birkdale': 'BDE',
        'Wellington Point': 'WPT',
        'Ormiston': 'ORO',
        'Cleveland': 'CVN',
        'Elimbah': 'EMH',
        'Beerburrum': 'BEB',
        'Glasshouse Mountains': 'GSS',
        'Beerwah': 'BWH',
        'Landsborough': 'LSH',
        'Mooloolah': 'MOH',
        'Eudlo': 'EUD',
        'Palmwoods': 'PAL',
        'Woombye': 'WOB',
        'Nambour': 'NBR',
        'Mayne Junction': 'MYJ',
        'Windsor': 'WID',
        'Wilston': 'WLQ',
        'Newmarket': 'NWM',
        'Alderley': 'ADY',
        'Enoggera': 'EGG',
        'Gaythorne': 'GAO',
        'Mitchelton': 'MHQ',
        'Oxford Park': 'OXP',
        'Grovely': 'GOQ',
        'Keperra': 'KEP',
        'Ferny Grove': 'FYG',
        'Robina Stabling Yard': 'ROBS',
        'Robina': 'ROB',
        'Varsity Lakes': 'VYS',
        'Caboolture West Yard': 'CAW',
        'International Airport': 'BIT',
        'Domestic Airport': 'BDT',
        'Ormeau': 'ORM',
        'Coomera': 'CXM',
        'Helensvale': 'HLN',
        'Nerang': 'NRG',
        'Beenleigh Stabling Yard': 'BNHS',
        'Beenleigh Stable': 'BNHS',
        'Banyo Stabling Yard': 'BQYS',
        'Wulkuraka Service Centre East': 'WFE',
        'WSC East Entrance': 'FEE',
        'Redbank Stabling Yard': 'RDKS',
        'Redbank Stabling':'RDKS',
        'Roma St Fork': 'RSF',
        'Clayfield': 'CYF',
        'Hendra': 'HDR',
        'Ascot': 'ACO',
        'Doomben': 'DBN',
        'Clapham Yard': 'CPM',
        'Varsity Lakes Turnback': 'VYST',
        'Varsity Lakes TB': 'VYST',
        'Woombye Stabling Yard': 'WOBS',
        'Gympie North': 'GYN',
        'Glanmire': 'GMR',
        'Woondum': 'WOO',
        'Traveston': 'TRA',
        'Cooran': 'COZ',
        'Pomona': 'PMQ',
        'Cooroy': 'COO',
        'Sunrise': 'SSE',
        'Eumundi': 'EUM',
        'North Arm': 'NHR',
        'Yandina': 'YAN',
        'Wulkuraka Service Centre West': 'WFW',
        'WSC West Entrance': 'FWE',
        'Tennyson': 'TNY',
        'Moolabin': 'MBN',
        'Rocklea sidings': 'RKET',
        'Rocklea Sidings': 'RKET',
        'Electric Train Balloon': 'ETB',
        'Elec Train Balloon': 'ETB',
        'Petrie Stabling Yard': 'PETS',
        'Petrie Eastern Sdgs': 'PETS',
        'Mayne East Stabling Yard':'MES',
        'Mayne North Stabling':'MNS',
        'Mayne 2':'MNE2',
        'Ormeau Stabling':'ORMS',
        'Pimpama':'PIA',
        'Hope Island':'HID',
        'Merrimac':'MRC',
        
        'Boggo Road':'BOG',
        'Boggo Road station':'BOG',
        'Albert Street':'ALB',
        'Woolloongabba':'WLG',
        
        'Mayne North Stabling':'MNS',
        'Mayne East':'MES',
        'Mayne East Stabling':'MES',
        'Clapham Yard':'CPM',
    
        'Comes From': 'CF',
        'Continues To': 'CT',
      }
    
    
        
    
    start_time = time.time()    
        
    
    
    mth_data = []
    fri_data = []
    sat_data = []
    sun_data = []
    
    
    data_dict = {
        mth_ws:mth_data,
        fri_ws:fri_data,
        sat_ws:sat_data,
        sun_ws:sun_data
        }
    
    
    
    for ws in [mth_ws,fri_ws,sat_ws,sun_ws]:
        
        
        traininfo = []
        tn_prev = ''
        
        
        
        for i,x in enumerate(tuple(ws.rows)):
            data = data_dict.get(ws)
            
            
            # if 0 < i < 3000:
            if 0 < i:
                a = [y.value for y in x]
                tn,stationnum,station,stoptype,line,peak,inout,direction,bnctime,bnchour,arr,dep,seats,cap,platform,load,util,designutil,board,alight = a
                
                # print(station)
                # stationName = stationmaster_keys[ stationmaster_vals.index(station) ]
                # traininfo = station
                
                infolist = [station,arr,dep,board,alight]
            
                #Add station
                if tn == tn_prev or i == 1:
                    # print(tn)
                    
                    # traininfo.append([board,station])
                    traininfo.append(infolist)
                    
                #First station 
                #Append previous info to data, initialise new list
                else:
                    data.append([tn_prev,traininfo])
                    traininfo = [infolist]
                    
                tn_prev = tn
        
        
        
            
            
            
            
            
            
    citystations = ['BHI','BRC','BNC','RS']
            
      
            
    filedate = datetime.now().strftime("%d-%b-%Y")
    filename_xlsx = f'ClosureImpacts-{filedate}.xlsx'   
    workbook = xlsxwriter.Workbook(filename_xlsx) 
    direct = workbook.add_worksheet('Impacts Summary')
    format1 = workbook.add_worksheet('Affected trips')
    format2 = workbook.add_worksheet('Affected trips (expanded)')
    
    topbottom = workbook.add_format({'bottom':1,'top':1})  
    mergedbold = workbook.add_format({'bold': True, 'align':'center','valign':'vcenter'})  
    # mergedboldb = workbook.add_format({'bold': True, 'align':'center','valign':'vcenter', 'bottom':1})    
    bold = workbook.add_format({'bold': True, 'align':'center'})   
    boldright = workbook.add_format({'bold': True, 'align':'right'})  
    center = workbook.add_format({'align':'center'}) 
    percentage = workbook.add_format({'num_format':9,'align':'center'})
    boldpillars = workbook.add_format({'align':'center','left':1, 'right':1,'bold':True})
    pillars = workbook.add_format({'align':'center','left':1, 'right':1})
    allbord = workbook.add_format({'align':'center','left':1, 'right':1, 'top':1, 'bottom':1})
    greyb = workbook.add_format({'bold': True, 'align':'center','border':1, 'bg_color':'#C0C0C0'})
    greytb = workbook.add_format({'bold': True, 'align':'center','border':2, 'bg_color':'#C0C0C0'})
    # workbook.add_format({'bold': True, 'align':'center','left':2, 'right':2, 'bg_color':'#C0C0C0'})
    
    
    
    
    
    ampeak_srt = '06:00:00'
    ampeak_end = '09:00:00'
    pmpeak_srt = '15:30:00'
    pmpeak_end = '18:30:00'
    
    mth_indirect = 0
    fri_indirect = 0
    sat_indirect = 0
    sun_indirect = 0
    
    
    mth_direct = [0,0]
    fri_direct = [0,0]
    sat_direct = [0,0]
    sun_direct = [0,0]
    
    mth_prepeak = [0,0]
    mth_ampeak  = [0,0]
    mth_intapeak = [0,0]
    mth_pmpeak  = [0,0]
    mth_postpeak = [0,0]
    fri_prepeak = [0,0]
    fri_ampeak  = [0,0]
    fri_intapeak = [0,0]
    fri_pmpeak  = [0,0]
    fri_postpeak = [0,0]
    sat_in  = [0,0]
    sat_out = [0,0]
    sun_in  = [0,0]
    sun_out = [0,0]
    
    
    for sheet in [format1,format2]:
        sheet.write(0,0,'Closures')
        sheet.write(0,1,', '.join(closures))
        sheet.write(2,1,'TrainID')
        # sheet.write(2,2,'Departure')
        sheet.write(2,3,'Stops')
        sheet.set_column(2,2,9)
    

    
    
    for dayidx,data in enumerate([mth_data,fri_data,sat_data,sun_data]):

        data_write = []
        rowf1 = 3
        rowf2 = 3
        for x in data:
            tn = x[0]
            # board = x[1]
            stops = x[1]
            stations = [x[0] for x in stops]
            oID,dID = stations[0],stations[-1]
            
            
            impacts = set(closures).intersection(set(stations))
            impactsidx = [stations.index(x) if x in stations else '' for x in impacts ]
            impactsidx.sort()
            
            
            thrucbd = set(citystations).intersection(set(stations))
            cbdidx = [stations.index(x) if x in stations else '' for x in thrucbd ]
            cbdidx.sort()
            
            if set(closures).intersection(stations):
                
                if 'BNC' not in stations:
                    print(i,stations)
                bncidx = stations.index('BNC')
                cbdtimingp = str(stops[bncidx][1])
                cbdarr = str(stops[bncidx][1])
                cbddep = str(stops[bncidx][2])

                #Option1
                format1.write(rowf1,1,tn)
                format1.write(rowf1,3,', '.join(stations))
                rowf1+=1
                
                #Option2
                format2.write(rowf2,1,tn)
                    
                for i,val in enumerate(stops):
                    # print(val[0],val[1],str(val[1]))
                    sID = val[0]
                    arr = str(val[1])
                    dep = str(val[2])
                    board = val[3]
                    alight = val[4]
                    
                    if i < min(impactsidx) < max(cbdidx):
                        ipsg = 2*(board - alight)
                        
                        if data == mth_data:
                            mth_indirect += ipsg
                        elif data == fri_data:
                            fri_indirect += ipsg
                        elif data == sat_data:
                            sat_indirect += ipsg
                        elif data == sun_data:
                            sun_indirect += ipsg
                    
                    font = bold if sID in closures else center
                    
                    
                    format2.write(rowf2,2,sID,font)
                    format2.write(rowf2,3,dep,font)
                    format2.write(rowf2,4,board,font)
                    
                    
                    city = i in cbdidx
                    inbound  =  i < min(cbdidx)
                    outbound =  max(cbdidx) < i
                    
                    
                    if sID in closures:
                    
                        for p,psg in enumerate([board,alight]):
                            # boardbool = False
                            # # boardbool = True
                            # psg = board if boardbool else alight
                            
                            
                            
                            
                            
                            
                            
                            if data == mth_data:
                                if cbdtimingp < ampeak_srt:
                                    mth_prepeak[p] += psg
                                elif ampeak_srt <= cbdtimingp < ampeak_end:
                                    mth_ampeak[p] += psg
                                elif ampeak_end <= cbdtimingp < pmpeak_srt:
                                    mth_intapeak[p] += psg
                                elif pmpeak_srt <= cbdtimingp < pmpeak_end:
                                    mth_pmpeak[p] += psg
                                elif pmpeak_end < cbdtimingp:
                                    mth_postpeak[p] += psg
                            elif data == fri_data:
                                if cbdtimingp < ampeak_srt:
                                    fri_prepeak[p] += psg
                                elif ampeak_srt <= cbdtimingp < ampeak_end:
                                    fri_ampeak[p] += psg
                                elif ampeak_end <= cbdtimingp < pmpeak_srt:
                                    fri_intapeak[p] += psg
                                elif pmpeak_srt <= cbdtimingp < pmpeak_end:
                                    fri_pmpeak[p] += psg
                                elif pmpeak_end < cbdtimingp:
                                    fri_postpeak[p] += psg
                                    
                                    
                                    
                            elif data == sat_data:
                                if city:
                                    pass
                                elif inbound:
                                    sat_in[p] += psg
                                elif outbound:
                                    sat_out[p] += psg
                                else:
                                    print('Investigate error')
                            elif data == sun_data:
                                if city:
                                    pass
                                elif inbound:
                                    sun_in[p] += psg
                                elif outbound:
                                    sun_out[p] += psg
                                else:
                                    # pass
                                    print('Investigate error')
    
                    rowf2+=1
                rowf2+=1

    
    
    
    direct.write(1,1,'Start Date',bold)
    direct.write(1,2,'End Date',bold)
    direct.write(2,1,sdate.strftime("%d-%b"),center)
    direct.write(2,2,edate_inclusive.strftime("%d-%b"),center)
    direct.write(1,4,'Closures:',boldright)
    direct.write(1,5,', '.join(closures))
    
    
    
    
    
    
    
    
    
    
    start_col = 4
    start_row = 4
    start_row2 = start_row + 30
    start_row3 = start_row2 + 26
    
    
    
    
    
    direct.merge_range( start_row+2,  start_col,   start_row+26, start_col,  'Direct impacts (weekday)', mergedbold)
    direct.merge_range( start_row+2,  start_col+1, start_row+7,  start_col+1,'Board',                    mergedbold)
    direct.merge_range( start_row+9,  start_col+1, start_row+13, start_col+1,'Board %',                  mergedbold)
    direct.merge_range( start_row+15, start_col+1, start_row+20, start_col+1,'Alight',                   mergedbold)
    direct.merge_range( start_row+22, start_col+1, start_row+26, start_col+1,'Alight %',                 mergedbold)
    
    periods = ['Pre-Peak','AM-Peak','Inter-Peak','PM-Peak','Post-Peak']
    direct.write_column( start_row+2,  start_col+2, periods,bold)
    direct.write(        start_row+7,  start_col+2,'Total', bold)
    direct.write_column( start_row+9,  start_col+2, periods,bold)
    direct.write_column( start_row+15, start_col+2, periods,bold)
    direct.write(        start_row+20, start_col+2,'Total', bold)
    direct.write_column( start_row+22, start_col+2, periods,bold)
    
    direct.merge_range(start_row2,start_col,start_row2+12,start_col,'Direct impacts (weekend)',mergedbold)
    direct.merge_range(start_row2,start_col+1,start_row2+2,start_col+1,'Board',mergedbold)
    direct.merge_range(start_row2+4,start_col+1,start_row2+5,start_col+1,'Board %',mergedbold)
    direct.merge_range(start_row2+7,start_col+1,start_row2+9,start_col+1,'Alight',mergedbold)
    direct.merge_range(start_row2+11,start_col+1,start_row2+12,start_col+1,'Alight %',mergedbold)
    
    
    
    
    
    in_out  = ['Inbound','Outbound']
    direct.write_column( start_row2,start_col+2,in_out,bold)
    direct.write(        start_row2+2,start_col+2,'Total',bold)
    direct.write_column( start_row2+4, start_col+2, in_out,bold)
    direct.write_column( start_row2+7, start_col+2, in_out,bold)
    direct.write(        start_row2+9, start_col+2, 'Total',bold)
    direct.write_column( start_row2+11, start_col+2, in_out,bold)
    
    
    direct.merge_range(  start_row3,   start_col,   start_row3+1, start_col,   'Indirect Impacts', mergedbold)
    direct.merge_range(  start_row3,   start_col+1, start_row3+1, start_col+1, 'Trips',            mergedbold)
    direct.write(        start_row3,   start_col+2,                            'Weekday',          bold)
    direct.write(        start_row3+1, start_col+2,                            'Weekend',          bold)
    
    
    
    
    
    
    mth_board = [mth_prepeak[0],mth_ampeak[0],mth_intapeak[0],mth_pmpeak[0],mth_postpeak[0]]
    fri_board = [fri_prepeak[0],fri_ampeak[0],fri_intapeak[0],fri_pmpeak[0],fri_postpeak[0]] 
    sat_board = [sat_in[0],sat_out[0]]
    sun_board = [sun_in[0],sun_out[0]]
    
    mth_alight = [mth_prepeak[1],mth_ampeak[1],mth_intapeak[1],mth_pmpeak[1],mth_postpeak[1]]
    fri_alight = [fri_prepeak[1],fri_ampeak[1],fri_intapeak[1],fri_pmpeak[1],fri_postpeak[1]]
    sat_alight = [sat_in[1],sat_out[1]]
    sun_alight = [sun_in[1],sun_out[1]]
    
    
    
    pre_peak_board  = 0
    am_peak_board   = 0
    inta_peak_board = 0
    pm_peak_board   = 0
    post_peak_board = 0
    
    # pre_peak_boardperc  = 0
    # am_peak_boardperc   = 0
    # inta_peak_boardperc = 0
    # pm_peak_boardperc   = 0
    # post_peak_boardperc = 0
    
    pre_peak_alight  = 0
    am_peak_alight   = 0
    inta_peak_alight = 0
    pm_peak_alight   = 0
    post_peak_alight = 0
    
    # pre_peak_alightperc  = 0
    # am_peak_alightperc   = 0
    # inta_peak_alightperc = 0
    # pm_peak_alightperc   = 0
    # post_peak_alightperc = 0
    
    
    in_board   = 0
    out_board  = 0
    in_alight  = 0
    out_alight = 0
    
    indirect_wkd  = 0
    indirect_wknd = 0
    
    
    
    x = [sdate+timedelta(days=x) for x in range(((edate)-sdate).days)]
    for i,y in enumerate(x,start_col+3):
        date = y.strftime("%d-%b")
        day = y.strftime("%A")
        direct.write(start_row,i,date,bold)
        direct.write(start_row+1,i,day,bold)
        
        
        
        
        
    
        if day in ['Monday','Tuesday','Wednesday','Thursday']:
            b_data = mth_board
            a_data = mth_alight
        elif day == 'Friday':
            b_data = fri_board
            a_data = fri_alight
        elif day == 'Saturday':
            b_data = sat_board
            a_data = sat_alight
        elif day == 'Sunday':
            b_data = sun_board
            a_data = sun_alight
        else:
            print('error')
    
        b_perc = [i/sum(b_data) if sum(b_data) else '' for i in b_data]
        a_perc = [i/sum(a_data) if sum(a_data) else '' for i in a_data]
        
        if day in ['Monday','Tuesday','Wednesday','Thursday','Friday']:
            pre_peak_board  += b_data[0]
            am_peak_board   += b_data[1]
            inta_peak_board += b_data[2]
            pm_peak_board   += b_data[3]
            post_peak_board += b_data[4]
            
            pre_peak_alight  += a_data[0]
            am_peak_alight   += a_data[1]
            inta_peak_alight += a_data[2]
            pm_peak_alight   += a_data[3]
            post_peak_alight += a_data[4]
            
    
            direct.write_column( start_row+2,i,b_data,       center)
            direct.write(        start_row+7,i,sum(b_data),  center)
            direct.write_column( start_row+9,i,b_perc,       percentage)
            direct.write_column( start_row+15,i,a_data,      center)
            direct.write(        start_row+20,i,sum(a_data), center)
            direct.write_column( start_row+22,i,a_perc,      percentage)
        
        
        elif day in ['Saturday','Sunday']:
            in_board   += b_data[0]
            out_board  += b_data[1]
            in_alight  += a_data[0]
            out_alight += a_data[1]
            
            direct.write_column( start_row2,  i,b_data,      center)
            direct.write(        start_row2+2,i,sum(b_data), center)
            direct.write_column( start_row2+4,i,b_perc,      percentage)
            direct.write_column( start_row2+7,i,a_data,      center)
            direct.write(        start_row2+9,i,sum(a_data), center)
            direct.write_column( start_row2+11,i,a_perc,      percentage)
            
            
        if day in ['Monday','Tuesday','Wednesday','Thursday']:
            direct.write(start_row3,i,mth_indirect,center)
            indirect_wkd += mth_indirect
        elif day == 'Friday':
            direct.write(start_row3,i,fri_indirect,center)
            indirect_wkd += fri_indirect
        elif day == 'Saturday':
            direct.write(start_row3+1,i,sat_indirect,center)
            indirect_wknd += sat_indirect
        elif day == 'Sunday':
            direct.write(start_row3+1,i,sun_indirect,center)
            indirect_wknd += sun_indirect
            
        
        
        
    
    
    totals_col      = len(x)+start_col+3
    weekly_board    = [pre_peak_board,am_peak_board,inta_peak_board,pm_peak_board,post_peak_board]
    weekly_alight   = [pre_peak_alight,am_peak_alight,inta_peak_alight,pm_peak_alight,post_peak_alight]
    
    weekend_board   = [in_board,out_board]
    weekend_alight  = [in_alight,out_alight]
    
    sum_board       = sum(weekly_board)
    sum_alight      = sum(weekly_alight)
    sum_week        = sum_board + sum_alight
    sum_wknd_board  = sum(weekend_board)
    sum_wknd_alight = sum(weekend_alight)
    sum_wknd        = sum_wknd_board + sum_wknd_alight
    
    total_sum       = sum_week + sum_wknd
    
    direct.write(       start_row+1,    totals_col,'Total',boldpillars)
    
    direct.write_column( start_row+2,   totals_col, weekly_board,    pillars)
    direct.write(        start_row+7,   totals_col, sum_board,       pillars)
    direct.write_column( start_row+8,   totals_col, 7*[''],          pillars)
    direct.write_column( start_row+15,  totals_col, weekly_alight,   pillars)
    direct.write(        start_row+20,  totals_col, sum_alight,      pillars)
    direct.write_column( start_row+21,  totals_col, 6*[''],          pillars)
    direct.write(        start_row+27,  totals_col, sum_week,        allbord)
    
            
    direct.write_column( start_row2,    totals_col, weekend_board,   pillars)
    direct.write(        start_row2+2,  totals_col, sum_wknd_board,  pillars)
    direct.write_column( start_row2+3,  totals_col, 4*[''],          pillars)
    direct.write_column( start_row2+7,  totals_col, weekend_alight,  pillars)    
    direct.write(        start_row2+9,  totals_col, sum_wknd_alight, pillars)  
    direct.write_column( start_row2+10, totals_col, 3*[''],          pillars)   
    direct.write(        start_row2+13, totals_col, sum_wknd,        allbord)
    
    direct.write(        start_row2+15, totals_col, total_sum,       greyb)
    
    
    
    
    total_indirect = indirect_wkd + indirect_wknd
    
    
    
    direct.write(start_row3,   totals_col, indirect_wkd,   pillars)
    direct.write(start_row3+1, totals_col, indirect_wknd,  pillars)
    direct.write(start_row3+2, totals_col, total_indirect, allbord)
    
    indirectanddirect = total_indirect + total_sum
    
    direct.write(start_row3+5, totals_col, indirectanddirect, greytb )
    
    
    
    
    direct.write_row( start_row+27,  start_col, 12*[''], topbottom)
    direct.write_row( start_row2+13, start_col, 12*[''], topbottom) 
    direct.write_row( start_row3+2,  start_col, 12*[''], topbottom)        

    
    direct.set_column(   start_col,    start_col,    30)
    direct.set_column(   start_col+2,  totals_col-1, 10.5)
            
    
    direct.conditional_format(start_row+9,start_col+3,start_row+13,start_col+11, {'type': '2_color_scale','min_color':'#FFFFFF','max_color':'#00B050','min_type': 'num'})
    direct.conditional_format(start_row+22,start_col+3,start_row+26,start_col+11, {'type': '2_color_scale','min_color':'#FFFFFF','max_color':'#00B050','min_type': 'num'})
    
    direct.conditional_format(start_row2+4,start_col+3,start_row2+5,start_col+11, {'type': '2_color_scale','min_color':'#FFFFFF','max_color':'#00B050','min_type': 'percent','min_type': 'num'})
    direct.conditional_format(start_row2+11,start_col+3,start_row2+12,start_col+11, {'type': '2_color_scale','min_color':'#FFFFFF','max_color':'#00B050','min_type': 'percent','min_type': 'num'})
    
    
    
    direct.activate()
    if OpenWorkbook:
        workbook.close()
        os.startfile(rf'{filename_xlsx}')    
    
        
    print(f'\n(runtime: {time.time()-start_time:.2f}seconds)')
    if ProcessDoneMessagebox:
        from tkinter import messagebox
        mb = tk.Tk()
        mb.withdraw()
        messagebox.showinfo('Closure Impacts','Process Done')
        
except Exception as e:
    logging.error(traceback.format_exc())
    if ProcessDoneMessagebox:
        time.sleep(15)