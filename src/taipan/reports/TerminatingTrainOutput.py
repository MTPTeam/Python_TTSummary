
from taipan.gui.base import select_file, show_info_scroll_safe
from PyQt6.QtWidgets import QApplication
import sys
import re
import os
import time
import logging
import traceback
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIG
# ============================================================
OUTPUT_NAME_SUFFIX = "TerminatingTrainOutput-"
AUTO_OPEN_AFTER_SAVE = True

HEADER_FILL_HEX = "BFBFBF"
HEADER_HEIGHT = 30
ORG_DEST_WIDTH = 30
SEPARATOR_GREY_HEX = "D3D3D3"
DWELL_NUMBER_FORMAT = "[m]:ss"

VALID_TRAIN_PAIRS = {
    "D": "E",
    "T": "A",
    "1": "2",
    "X": "W",
    "R": "N"
}

ALLOWED_TRAIN_PREFIXES = set(VALID_TRAIN_PAIRS.keys()) | set(VALID_TRAIN_PAIRS.values())

weekdaykey_dict = {
    '120': 'Mon-Thu',
    '64':  'Mon',
    '32':  'Tue',
    '16':  'Wed',
    '8':   'Thu',
    '4':   'Fri',
    '2':   'Sat',
    '1':   'Sun'
}

# ============================================================
# HELPERS
# ============================================================
def timetrim_hm(timestring):
    if timestring is None:
        return None
    s = str(timestring).strip()
    if ':' not in s:
        return s
    s = s[:-3]
    if s.startswith("0") and len(s) > 1 and s[1].isdigit():
        s = s[1:]
    return s

def time_to_seconds_hms(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    s = str(x).strip()
    parts = s.split(":")
    if len(parts) != 3:
        return None
    h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
    return h * 3600 + m * 60 + sec

def excel_duration(seconds):
    return float(seconds) / 86400.0

def format_run(run):
    if '-' in run:
        return run
    newrun = ''
    if run.startswith("E") and len(run) > 1 and run[1].isnumeric():
        digits = ''.join(x for x in run[1:] if x.isnumeric())
        letters = ''.join(x for x in run[1:] if x.isalpha())
        newrun = f"E{digits}-{letters}" if run.endswith(('A', 'B')) else f"E{digits}"
    elif len(run) >= 3 and run[0].isnumeric() and run[2] == "E":
        digits = ''.join(x for x in run if x.isnumeric())
        letters = ''.join(x for x in run[3:] if x.isalpha())
        newrun = f"{digits}E-{letters}" if run.endswith(('A', 'B')) else f"{digits}E"
    elif run[0].isnumeric() and run.endswith(('A', 'B')):
        digits = ''.join(x for x in run if x.isnumeric())
        letters = ''.join(x for x in run if x.isalpha())
        newrun = f"{digits}-{letters}"
    elif run[0].isalpha() and run.endswith(('1', '2')):
        letters = ''.join(x for x in run if x.isalpha())
        digits = ''.join(x for x in run if x.isnumeric())
        newrun = f"{letters}-{digits}"
    return newrun if newrun else run

def to_yard_flag(dest_text):
    if dest_text is None:
        return "No"
    s = str(dest_text).strip()
    if not s:
        return "No"
    if s.lower().endswith("gympie north"):
        return "No"
    last_word = s.split()[-1].lower()
    yard_words = {"south", "flyover", "west", "east", "north", "yard"}
    return "Yes" if last_word in yard_words else "No"

def is_yard_like_station(station_name):
    if station_name is None:
        return False
    s = str(station_name).strip().lower()
    return any(term in s for term in ["yard", "stabling", "depot"])

def format_risk_train_list(df, tids_col="TID", sort_col="Start_raw"):
    if df.empty:
        return ""
    if sort_col in df.columns:
        df = df.sort_values(sort_col, kind="mergesort")
    tids = df[tids_col].dropna().astype(str).tolist()
    if not tids:
        return ""
    preview = tids[:5]
    text = ", ".join(preview)
    if len(tids) > 5:
        text += " ..."
    return f"{len(tids)}: {text}"

def resolve_terminating_station_and_dwell(current_station, current_dwell, fallback_station, fallback_dwell):
    try:
        dwell_int = int(float(current_dwell)) if current_dwell not in (None, "") else None
    except ValueError:
        dwell_int = None
    if is_yard_like_station(current_station) and dwell_int == 1:
        if fallback_station not in (None, "") and fallback_dwell not in (None, ""):
            return fallback_station, fallback_dwell
    return current_station, current_dwell

def get_standalone_dwell_from_entries(entries):
    ignore_station_names = {"Campbell St", "Exhibition", "Normanby", "Mayne Yard Arrival", "Mayne"}
    for entry in reversed(entries):
        attrib = entry.attrib
        stop_time = attrib.get("stopTime")
        track_id = attrib.get("trackID")
        station_name = attrib.get("stationName", "").strip()
        platform = None
        if track_id:
            match = re.search(r"\d+", track_id)
            if match:
                platform = f"P-{match.group()}"
        if stop_time is None or track_id is None:
            continue
        if not (track_id.startswith("D") or track_id.startswith("U")):
            continue
        if station_name in ignore_station_names:
            continue
        try:
            stop_time_int = int(float(stop_time))
        except ValueError:
            continue
        return stop_time_int, station_name, platform
    return None, None, None

def get_extra_dwell_text_from_entries(entries):
    if not entries or len(entries) < 3:
        return "No"
    matches = []
    for entry in entries[1:-1]:
        attrib = entry.attrib
        stop_time = attrib.get("stopTime")
        track_id = attrib.get("trackID")
        station_name = attrib.get("stationName", "").strip()
        if stop_time is None or track_id is None:
            continue
        if not (track_id.startswith("D") or track_id.startswith("U")):
            continue
        try:
            stop_time_int = int(float(stop_time))
        except ValueError:
            continue
        if stop_time_int <= 1:
            continue
        if station_name:
            matches.append(f"{stop_time_int}s @ {station_name}")
    return " | ".join(matches) if matches else "No"

def get_last_valid_summary_stop(entries):
    ignore_station_names = {"Campbell St", "Exhibition", "Normanby", "Mayne Yard Arrival", "Mayne"}
    for entry in reversed(entries):
        attrib = entry.attrib
        stop_time = attrib.get("stopTime")
        track_id = attrib.get("trackID")
        station_name = attrib.get("stationName", "").strip()
        if stop_time is None or track_id is None:
            continue
        if not (track_id.startswith("D") or track_id.startswith("U")):
            continue
        if station_name in ignore_station_names:
            continue
        try:
            stop_time_int = int(float(stop_time))
        except ValueError:
            continue
        if stop_time_int <= 1:
            continue
        return station_name, stop_time_int
    return None, None

# ============================================================
# CORE: parse RSX and build nursery dataframe
# ============================================================
def build_nursery_rows_from_rsx(path):
    tree = ET.parse(path)
    root = tree.getroot()

    excluded_count = 0
    excluded_prefixes = set()
    missing_run_count = 0
    missing_run_tids = []

    tn_list = []
    tn_doubles = []

    for train in root.iter('train'):
        tn = train.attrib.get('number')
        day = train[0][0][0].attrib.get('weekdayKey')

        if tn:
            first_char = str(tn).strip()[0]
            if first_char not in ALLOWED_TRAIN_PREFIXES:
                excluded_count += 1
                excluded_prefixes.add(first_char)
                continue

        run = train.attrib.get("lineID")
        if not run or str(run).strip() == "":
            missing_run_count += 1
            missing_run_tids.append(tn)
            continue

        if (tn, day) in tn_list:
            tn_doubles.append((tn, day))
        tn_list.append((tn, day))

    if tn_doubles:
        msg = "Error: Duplicate train numbers detected:\n\n"
        for tn, day in tn_doubles:
            msg += f"- {weekdaykey_dict.get(day, day)} has duplicate train number {tn}\n"
        raise ValueError(msg)

    if missing_run_count > 0:
        sample = ", ".join(str(t) for t in missing_run_tids)
        show_info_scroll_safe(
            "Missing Run Error",
            f"{missing_run_count} valid train(s) are missing a lineID (Run).\n\n"
            f"Affected Services:\n{sample}\n\n"
            f"This is a critical data issue. Please assign lineID values and re-run the tool."
        )
        raise ValueError(f"Missing a lineID (Run).")

    nursery_rows = []

    for train in root.iter('train'):
        tn = train.attrib.get('number')
        WeekdayKey = train[0][0][0].attrib.get('weekdayKey')

        if tn:
            first_char = str(tn).strip()[0]
            if first_char not in ALLOWED_TRAIN_PREFIXES:
                continue

        lineID = train.attrib.get('lineID')
        if not lineID or str(lineID).strip() == "":
            continue

        entries = [x for x in train.iter('entry')]
        if not entries:
            continue

        origin = entries[0].attrib
        destin = entries[-1].attrib

        standalone_dwell_sec, standalone_dwell_station, platform = get_standalone_dwell_from_entries(entries)
        extra_dwell_text = get_extra_dwell_text_from_entries(entries)
        summary_prev_station, summary_prev_dwell_sec = get_last_valid_summary_stop(entries)

        odep = origin.get('departure')
        ddep = destin.get('departure')
        org_name = origin.get('stationName')
        dest_name = destin.get('stationName')

        lineID = train.attrib.get('lineID', '')
        run = lineID.split('~', 1)[1][1:] if '~' in lineID else lineID
        run = format_run(run)

        base = {
            "Run": run,
            "TID": tn,
            "Start": timetrim_hm(odep),
            "Org": org_name,
            "Dest": dest_name,
            "Finish": timetrim_hm(ddep),
            "Start_raw": odep,
            "Finish_raw": ddep,
            "StandaloneDwellSec": standalone_dwell_sec,
            "StandaloneDwellStation": standalone_dwell_station,
            "ExtraDwellText": extra_dwell_text,
            "SummaryPrevStation": summary_prev_station,
            "SummaryPrevDwellSec": summary_prev_dwell_sec,
            "Platform": platform,
        }

        if WeekdayKey in ['64', '120']:
            r = dict(base); r["Day"] = "M______"; nursery_rows.append(r)
        if WeekdayKey in ['32', '120']:
            r = dict(base); r["Day"] = "_T_____"; nursery_rows.append(r)
        if WeekdayKey in ['16', '120']:
            r = dict(base); r["Day"] = "__W____"; nursery_rows.append(r)
        if WeekdayKey in ['8', '120']:
            r = dict(base); r["Day"] = "___T___"; nursery_rows.append(r)
        if WeekdayKey in ['4']:
            r = dict(base); r["Day"] = "____F__"; nursery_rows.append(r)
        if WeekdayKey in ['2']:
            r = dict(base); r["Day"] = "_____S_"; nursery_rows.append(r)
        if WeekdayKey in ['1']:
            r = dict(base); r["Day"] = "______S"; nursery_rows.append(r)

    df = pd.DataFrame(nursery_rows)
    df["Start_sort"] = df["Start_raw"].apply(time_to_seconds_hms)
    df = df.sort_values(["Day", "Run", "Start_sort"], kind="mergesort").reset_index(drop=True)

    if excluded_count > 0:
        prefix_list = ", ".join(sorted(excluded_prefixes))
        show_info_scroll_safe(
            "QA Notice: Excluded Services",
            f"{excluded_count} service(s) were excluded based on unsupported train number prefixes.\n\n"
            f"Prefixes detected:\n{prefix_list}\n\n"
            f"This is expected behaviour.\n\n"
            f"Please wait momentarily while the Excel file opens automatically.\n\n"
            f"The associated PDF summary can be found in the same folder as the selected RSX file."
        )
    else:
        show_info_scroll_safe(
            "Data Integrity Check Passed",
            "No anomalies detected.\n\n"
            "Please wait momentarily while the Excel file opens automatically.\n\n"
            "The associated PDF summary can be found in the same folder as the selected RSX file."
        )

    return df, {"excluded_count": excluded_count, "excluded_prefixes": excluded_prefixes}


# ============================================================
# CORE: build final output rows (pairs, standalone, dwell, To Yard?)
# ============================================================
def build_final_output(df_nursery, day_filter=None):
    if day_filter is not None:
        if day_filter == "Mon-Thu":
            df_nursery = df_nursery[
                df_nursery["Day"].isin(["M______", "_T_____", "__W____", "___T___"])
            ].copy()
        elif day_filter == "Fri":
            df_nursery = df_nursery[df_nursery["Day"] == "____F__"].copy()
        elif day_filter == "Sat":
            df_nursery = df_nursery[df_nursery["Day"] == "_____S_"].copy()
        elif day_filter == "Sun":
            df_nursery = df_nursery[df_nursery["Day"] == "______S"].copy()

    df_nursery = df_nursery.sort_values(["Run", "Start_sort"], kind="mergesort").reset_index(drop=True)

    output_rows = []
    pair_id = 0

    def blank_row():
        return {
            "Run": None, "TID": None, "Day": None, "Start": None, "Org": None, "Dest": None, "Finish": None,
            "Platform": None, "Dwell": None, "To Yard?": None, "Additional Dwell (en route)": None,
            "__pair_id": None
        }

    for run, group in df_nursery.groupby("Run", sort=False):
        group = group.sort_values("Start_sort", kind="mergesort").reset_index(drop=True)

        i = 0
        while i < len(group) - 1:
            cur_tid = str(group.loc[i, "TID"])
            nxt_tid = str(group.loc[i + 1, "TID"])

            is_pair = (
                cur_tid and nxt_tid and
                VALID_TRAIN_PAIRS.get(cur_tid[0]) == nxt_tid[0]
            )

            if is_pair:
                pair_id += 1

                r1 = {
                    "Run": group.loc[i, "Run"],
                    "TID": group.loc[i, "TID"],
                    "Day": group.loc[i, "Day"],
                    "Start": group.loc[i, "Start"],
                    "Org": group.loc[i, "Org"],
                    "Dest": group.loc[i, "Dest"],
                    "Finish": group.loc[i, "Finish"],
                    "Platform": group.loc[i].get("Platform"),
                    "Dwell": None,
                    "To Yard?": None,
                    "Additional Dwell (en route)": None,
                    "__pair_id": pair_id,
                    "SummaryPrevStation": group.loc[i].get("SummaryPrevStation"),
                    "SummaryPrevDwellSec": group.loc[i].get("SummaryPrevDwellSec"),
                }

                r2 = {
                    "Run": group.loc[i + 1, "Run"],
                    "TID": group.loc[i + 1, "TID"],
                    "Day": group.loc[i + 1, "Day"],
                    "Start": group.loc[i + 1, "Start"],
                    "Org": group.loc[i + 1, "Org"],
                    "Dest": group.loc[i + 1, "Dest"],
                    "Finish": group.loc[i + 1, "Finish"],
                    "Platform": group.loc[i].get("Platform"),
                    "Dwell": None,
                    "To Yard?": None,
                    "Additional Dwell (en route)": None,
                    "__pair_id": pair_id,
                    "SummaryPrevStation": group.loc[i + 1].get("SummaryPrevStation"),
                    "SummaryPrevDwellSec": group.loc[i + 1].get("SummaryPrevDwellSec"),
                }

                s2 = time_to_seconds_hms(group.loc[i + 1, "Start_raw"])
                f1 = time_to_seconds_hms(group.loc[i, "Finish_raw"])
                dwell_sec = (s2 - f1) + 1

                while dwell_sec < 0:
                    dwell_sec += 24 * 3600

                term_station, resolved_dwell = resolve_terminating_station_and_dwell(
                    r1["Dest"],
                    dwell_sec,
                    group.loc[i].get("SummaryPrevStation"),
                    group.loc[i].get("SummaryPrevDwellSec")
                )

                r1["Terminating Station"] = term_station
                r2["Terminating Station"] = term_station
                r1["Dwell"] = resolved_dwell
                r2["Dwell"] = resolved_dwell

                yard_val = to_yard_flag(r2["Dest"])
                r1["To Yard?"] = yard_val
                r2["To Yard?"] = yard_val

                extra_dwell_val = group.loc[i + 1].get("ExtraDwellText", "No")
                r1["Additional Dwell (en route)"] = extra_dwell_val
                r2["Additional Dwell (en route)"] = extra_dwell_val

                output_rows.append(r1)
                output_rows.append(r2)
                output_rows.append(blank_row())

                i += 2
            else:
                i += 1

        last_row = group.loc[len(group) - 1]
        last_tid = str(last_row["TID"])
        if last_tid and last_tid[0] in VALID_TRAIN_PAIRS.keys():
            term_station, resolved_dwell = resolve_terminating_station_and_dwell(
                last_row.get("StandaloneDwellStation", None),
                last_row.get("StandaloneDwellSec", None),
                last_row.get("SummaryPrevStation"),
                last_row.get("SummaryPrevDwellSec"),
            )

            r = {
                "Run": last_row["Run"],
                "TID": last_row["TID"],
                "Day": last_row["Day"],
                "Start": last_row["Start"],
                "Org": last_row["Org"],
                "Dest": last_row["Dest"],
                "Finish": last_row["Finish"],
                "Platform": last_row.get("Platform"),
                "Dwell": resolved_dwell,
                "Terminating Station": term_station,
                "To Yard?": to_yard_flag(last_row["Dest"]),
                "Additional Dwell (en route)": "No",
                "__pair_id": None,
                "SummaryPrevStation": last_row.get("SummaryPrevStation"),
                "SummaryPrevDwellSec": last_row.get("SummaryPrevDwellSec"),
            }

            output_rows.append(r)
            output_rows.append(blank_row())

    output_df = pd.DataFrame(output_rows)

    visible_column_order = [
        "Run", "TID", "Day", "Start", "Org", "Dest", "Finish",
        "Terminating Station", "Dwell", "To Yard?", "Additional Dwell (en route)"
    ]

    helper_columns = ["Platform", "SummaryPrevStation", "SummaryPrevDwellSec", "__pair_id"]

    visible_column_order = [c for c in visible_column_order if c in output_df.columns]
    helper_columns = [c for c in helper_columns if c in output_df.columns]

    output_df = output_df[visible_column_order + helper_columns]

    return output_df


# ============================================================
# Summary Page
# ============================================================
def create_summary_sheet(wb, output_df, sheet_name="Summary"):
    ws = wb.create_sheet(title=sheet_name)

    day_label = sheet_name.replace("Summary (", "").replace(")", "")

    full_day_map = {
        "Mon-Thu": "Monday to Thursday",
        "Fri": "Friday",
        "Sat": "Saturday",
        "Sun": "Sunday"
    }

    day_label = full_day_map.get(day_label, day_label)

    df_real = output_df[output_df["Run"].notna()].copy()

    df_primary = df_real[
        df_real["__pair_id"].isna() | ~df_real["__pair_id"].duplicated()
    ].copy()

    df_primary["Dwell_sec"] = pd.to_numeric(df_primary["Dwell"], errors="coerce")
    df_primary["Dwell_min"] = df_primary["Dwell_sec"] / 60

    df_primary["SummaryTerminatingStation"] = df_primary["Terminating Station"]
    df_primary["SummaryDwellSec"] = df_primary["Dwell_sec"]

    mask_yard_artifact = (
        df_primary["SummaryDwellSec"].eq(1) &
        df_primary["SummaryTerminatingStation"].apply(is_yard_like_station)
    )

    if "SummaryPrevStation" in df_primary.columns and "SummaryPrevDwellSec" in df_primary.columns:
        df_primary.loc[mask_yard_artifact, "SummaryTerminatingStation"] = df_primary.loc[
            mask_yard_artifact, "SummaryPrevStation"
        ]
        df_primary.loc[mask_yard_artifact, "SummaryDwellSec"] = df_primary.loc[
            mask_yard_artifact, "SummaryPrevDwellSec"
        ]
    else:
        df_primary.loc[mask_yard_artifact, "SummaryDwellSec"] = None

    df_primary["SummaryDwellMin"] = df_primary["SummaryDwellSec"] / 60

    df_1s_non_yard = df_primary[
        (df_primary["Dwell_sec"] == 1) &
        (~df_primary["Terminating Station"].apply(is_yard_like_station))
    ].copy()

    one_sec_summary = (
        df_1s_non_yard.groupby("Terminating Station")
        .agg(
            Count=("Run", "count"),
            Platforms=("Platform", lambda x: sorted(
                set(p for p in x if pd.notna(p)),
                key=lambda p: int(str(p).split("-")[1]) if "-" in str(p) else 999
            ))
        )
        .reset_index()
        .sort_values("Count", ascending=False)
    )

    total_terminating_services = len(df_primary)
    services_direct_to_yard = (df_primary["To Yard?"] == "Yes").sum()

    df_to_yard = df_primary[
        (df_primary["To Yard?"] == "Yes") &
        (df_primary["SummaryDwellSec"].notna()) &
        (df_primary["SummaryDwellSec"] > 1)
    ].copy()

    if not df_to_yard.empty:
        shortest_val = df_to_yard["SummaryDwellSec"].min()
        count_shortest = len(df_to_yard[df_to_yard["SummaryDwellSec"] == shortest_val])
        shortest_dwell_to_yard_display = f"{int(shortest_val)}s x {count_shortest} train(s)"
        longest_val = df_to_yard["SummaryDwellSec"].max()
        count_longest = len(df_to_yard[df_to_yard["SummaryDwellSec"] == longest_val])
        longest_dwell_to_yard_display = f"{int(longest_val)}s x {count_longest} train(s)"
    else:
        shortest_dwell_to_yard_display = "N/A"
        longest_dwell_to_yard_display = "N/A"

    df_valid_dwell = df_primary[df_primary["SummaryDwellSec"] > 1].copy()

    if not df_valid_dwell.empty:
        min_val = df_valid_dwell["SummaryDwellSec"].min()
        stations = (
            df_valid_dwell[df_valid_dwell["SummaryDwellSec"] == min_val]
            ["SummaryTerminatingStation"].dropna().unique()
        )
        station_list = ", ".join(sorted(stations))
        shortest_dwell_label = f"Shortest Dwell: {int(min_val)}s @ {station_list}"
    else:
        shortest_dwell_label = "Shortest Dwell: No valid dwell data"

    station_counts = (
        df_primary[df_primary["SummaryTerminatingStation"].notna()]
        .groupby("SummaryTerminatingStation")
        .size()
        .reset_index(name="Terminating_Trains")
        .sort_values("Terminating_Trains", ascending=False)
    )

    if not station_counts.empty:
        top_station_row = station_counts.iloc[0]
        top_station = top_station_row["SummaryTerminatingStation"]
        top_station_count = int(top_station_row["Terminating_Trains"])
        most_active_label = f"Most Active Terminus: {top_station} ({top_station_count} trains)"
    else:
        most_active_label = "Most Active Terminus: No data"

    high_risk_services = len(df_primary[
        (df_primary["To Yard?"] == "Yes") &
        (df_primary["SummaryDwellSec"].notna()) &
        (df_primary["SummaryDwellSec"] <= 30)
    ])

    medium_risk_services = len(df_primary[
        (df_primary["To Yard?"] == "Yes") &
        (df_primary["SummaryDwellSec"].notna()) &
        (df_primary["SummaryDwellSec"] > 30) &
        (df_primary["SummaryDwellSec"] <= 90)
    ])

    low_risk_services = len(df_primary[
        (df_primary["To Yard?"] == "Yes") &
        (df_primary["SummaryDwellSec"].notna()) &
        (df_primary["SummaryDwellSec"] > 90) &
        (df_primary["SummaryDwellSec"] <= 150)
    ])

    minimum_risk_services = len(df_primary[
        (df_primary["To Yard?"] != "Yes") |
        (
            (df_primary["To Yard?"] == "Yes") &
            (df_primary["SummaryDwellSec"].notna()) &
            (df_primary["SummaryDwellSec"] > 150)
        )
    ])

    station_summary = (
        df_primary[df_primary["SummaryTerminatingStation"].notna()]
        .groupby("SummaryTerminatingStation")
        .agg(
            Terminating_Trains=("Run", "count"),
            Trains_to_Yard=("To Yard?", lambda s: int((s == "Yes").sum())),
        )
        .reset_index()
        .rename(columns={"SummaryTerminatingStation": "Terminating Station"})
    )

    min_dwell_to_yard = (
        df_primary[
            (df_primary["SummaryTerminatingStation"].notna()) &
            (df_primary["To Yard?"] == "Yes") &
            (df_primary["SummaryDwellSec"].notna()) &
            (df_primary["SummaryDwellSec"] > 1)
        ]
        .groupby("SummaryTerminatingStation")
        .agg(Min_Dwell_To_Yard_Sec=("SummaryDwellSec", "min"))
        .reset_index()
        .rename(columns={"SummaryTerminatingStation": "Terminating Station"})
    )

    station_summary = station_summary.merge(min_dwell_to_yard, on="Terminating Station", how="left")

    risk_rows = []
    for station, group in df_primary[df_primary["SummaryTerminatingStation"].notna()].groupby("SummaryTerminatingStation"):
        g = group.copy()

        high_df = g[(g["To Yard?"] == "Yes") & (g["SummaryDwellSec"].notna()) & (g["SummaryDwellSec"] <= 30)]
        medium_df = g[(g["To Yard?"] == "Yes") & (g["SummaryDwellSec"].notna()) & (g["SummaryDwellSec"] > 30) & (g["SummaryDwellSec"] <= 90)]
        low_df = g[(g["To Yard?"] == "Yes") & (g["SummaryDwellSec"].notna()) & (g["SummaryDwellSec"] > 90) & (g["SummaryDwellSec"] <= 150)]
        minimum_df = g[(g["To Yard?"] != "Yes") | ((g["To Yard?"] == "Yes") & (g["SummaryDwellSec"].notna()) & (g["SummaryDwellSec"] > 150))]

        risk_rows.append({
            "Terminating Station": station,
            "High Risk Count": len(high_df),
            "Medium Risk Count": len(medium_df),
            "Low Risk Count": len(low_df),
            "High Risk": format_risk_train_list(high_df),
            "Medium Risk": format_risk_train_list(medium_df),
            "Low Risk": format_risk_train_list(low_df),
            "Minimum Risk": format_risk_train_list(minimum_df),
        })

    risk_summary = pd.DataFrame(risk_rows, columns=[
        "Terminating Station", "High Risk Count", "Medium Risk Count", "Low Risk Count",
        "High Risk", "Medium Risk", "Low Risk", "Minimum Risk"
    ])

    station_summary = station_summary.merge(risk_summary, on="Terminating Station", how="left")

    for col in ["High Risk Count", "Medium Risk Count", "Low Risk Count"]:
        if col in station_summary.columns:
            station_summary[col] = station_summary[col].fillna(0).astype(int)

    for col in ["High Risk", "Medium Risk", "Low Risk", "Minimum Risk"]:
        if col in station_summary.columns:
            station_summary[col] = station_summary[col].fillna("")

    station_summary = station_summary.sort_values(
        ["High Risk Count", "Medium Risk Count", "Low Risk Count", "Terminating_Trains"],
        ascending=[False, False, False, False]
    )

    TOP_N = 10
    station_summary = station_summary.head(TOP_N)

    if not station_summary.empty:
        station_summary["Min_Dwell_To_Yard_Sec"] = station_summary["Min_Dwell_To_Yard_Sec"].fillna(0).astype(int)
        for col in ["High Risk", "Medium Risk", "Low Risk", "Minimum Risk"]:
            if col in station_summary.columns:
                station_summary[col] = station_summary[col].fillna("")

    # ── Colour palette ──────────────────────────────────────────────────────────
    # Maroon family
    MAROON_DARK    = "6B0000"   # title bar (deeper, richer)
    MAROON_MID     = "A61C1C"   # section header bars
    MAROON_LIGHT   = "C62828"   # tile value text (neutral KPIs)
    # Tile accent bars — saturated versions of each risk for contrast on white
    HIGH_ACCENT    = "C0392B"   # strong red-rose accent bar
    MEDIUM_ACCENT  = "CA6F1E"   # strong amber-orange accent bar
    LOW_ACCENT     = "B7950B"   # strong gold accent bar
    MINIMUM_ACCENT = "717D7E"   # medium slate-grey accent bar
    # Tile body risk fills — deliberately lighter / more muted than accent
    HIGH_C         = "F2D7D5"   # very soft rose
    MEDIUM_C       = "FAE5D3"   # very soft peach
    LOW_C          = "FDFBD4"   # very soft straw
    MINIMUM_C      = "E8EAED"   # cool blue-grey (distinct from table header grey)
    # Neutrals
    WHITE          = "FFFFFF"
    NEAR_WHITE     = "F9F9F9"
    TILE_SPACER    = "F2F2F2"   # slightly off-white for label row
    LIGHT_GREY     = "EDEDED"   # alt table rows
    TABLE_HDR_GREY = "CACACA"   # table header — clearly different from MINIMUM_C
    DARK_GREY      = "1A1A1A"
    SUBTLE_GREY    = "5A5A5A"
    WARNING_BG     = "F5F5F5"

    def _fill(hex_col):
        return PatternFill(start_color=hex_col, end_color=hex_col, fill_type="solid")

    title_fill      = _fill(MAROON_DARK)
    section_fill    = _fill(MAROON_MID)
    white_fill      = _fill(WHITE)
    near_white_fill = _fill(NEAR_WHITE)
    tile_spacer_fill= _fill(TILE_SPACER)
    table_hdr_fill  = _fill(TABLE_HDR_GREY)
    warning_bg_fill = _fill(WARNING_BG)
    alt_row_fill    = _fill(LIGHT_GREY)

    HIGH_RISK_FILL    = _fill(HIGH_C)
    MEDIUM_RISK_FILL  = _fill(MEDIUM_C)
    LOW_RISK_FILL     = _fill(LOW_C)
    MINIMUM_RISK_FILL = _fill(MINIMUM_C)

    # Fonts
    title_font        = Font(color=WHITE,        bold=True,  size=14, name="Calibri")
    section_font      = Font(color=WHITE,        bold=True,  size=9,  name="Calibri")
    body_font         = Font(color=DARK_GREY,                size=8,  name="Calibri")
    tile_value_font   = Font(color=MAROON_LIGHT, bold=True,  size=16, name="Calibri")
    tile_label_font   = Font(color=SUBTLE_GREY,              size=7,  name="Calibri")
    dark_value_font   = Font(color=DARK_GREY,    bold=True,  size=16, name="Calibri")
    table_header_font = Font(color=DARK_GREY,    bold=True,  size=8,  name="Calibri")
    warning_font      = Font(color="4A2700",                 size=8,  name="Calibri")
    ok_font           = Font(color="1B4F1B",                 size=8,  name="Calibri")
    legend_font       = Font(color=DARK_GREY,                size=8,  name="Calibri")
    legend_bold_font  = Font(color=DARK_GREY,    bold=True,  size=8,  name="Calibri")
    note_font         = Font(color="999999",     italic=True,size=7,  name="Calibri")
    station_font      = Font(color=DARK_GREY,    bold=True,  size=8,  name="Calibri")
    data_font         = Font(color=DARK_GREY,                size=8,  name="Calibri")

    # Borders
    thin_side   = Side(style="thin",   color="CCCCCC")
    medium_side = Side(style="medium", color="999999")
    maroon_side = Side(style="medium", color=MAROON_MID)
    no_side     = Side(style=None)

    thin_box    = Border(left=thin_side,   right=thin_side,   top=thin_side,   bottom=thin_side)
    warn_border = Border(left=maroon_side, right=thin_side,   top=thin_side,   bottom=thin_side)

    # Alignments
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    left_nowrap = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── Page setup ──────────────────────────────────────────────────────────────
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left   = 0.28
    ws.page_margins.right  = 0.28
    ws.page_margins.top    = 0.30
    ws.page_margins.bottom = 0.30
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.print_options.horizontalCentered = True

    for col in ["A","B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 18.0

    # ── Layout ──────────────────────────────────────────────────────────────────
    # Row  1    : Title bar
    # Row  2    : Objective section header
    # Row  3    : Objective text
    # Row  4    : white spacer (separates objective from KPI tiles)
    # Row  5    : KPI section bar
    # Row  6    : white spacer (so tile accent bars don't bleed into maroon above)
    # Rows 7-9  : KPI tile row 1  (accent / value / label)
    # Row  10   : spacer
    # Rows 11-13: KPI tile row 2  (accent / value / label)
    # Row  14   : spacer
    # Row  15   : DQ header (A–E) | Legend header (F–H)
    # Row  16+  : DQ rows (A–E, no wrap) | Legend rows (F–H)  ← dynamic
    # ...       : note row
    # ...       : Top-N table

    ws.row_dimensions[1].height  = 26   # title
    ws.row_dimensions[2].height  = 13   # Objective header
    ws.row_dimensions[3].height  = 18   # Objective text
    ws.row_dimensions[4].height  = 4    # spacer
    ws.row_dimensions[5].height  = 13   # KPI section bar
    ws.row_dimensions[6].height  = 4    # WHITE gap — tiles will visually "float" off the bar
    ws.row_dimensions[7].height  = 6    # tile accent bar row 1
    ws.row_dimensions[8].height  = 8    # tile label row 1  ← much shorter
    ws.row_dimensions[9].height  = 30   # tile value row 1  ← much taller
    ws.row_dimensions[10].height = 5    # spacer between tile rows
    ws.row_dimensions[11].height = 6    # tile accent bar row 2
    ws.row_dimensions[12].height = 8    # tile label row 2  ← much shorter
    ws.row_dimensions[13].height = 30   # tile value row 2  ← much taller
    ws.row_dimensions[14].height = 5    # spacer

    # ── ROW 1: Title ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = f"Terminating Service Dwell Time Summary  —  {day_label}"
    c.fill      = title_fill
    c.font      = title_font
    c.alignment = center

    # ── ROWS 2-3: Objective (first, before KPIs) ─────────────────────────────────
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = "Objective"
    c.fill      = section_fill
    c.font      = section_font
    c.alignment = left

    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = (
        f"This summary presents dwell time characteristics for terminating trains on {day_label}, "
        "with the objective of identifying higher risk stations and services to strengthen safe railway operations."
    )
    c.fill      = near_white_fill
    c.font      = body_font
    c.alignment = left
    c.border    = thin_box

    # ── ROW 5: KPI section bar ───────────────────────────────────────────────────
    ws.merge_cells("A5:H5")
    c = ws["A5"]
    c.value     = "Key Performance Indicators"
    c.fill      = section_fill
    c.font      = section_font
    c.alignment = left

    # Row 6: white gap so tile accent bars are clearly visible below the maroon bar
    ws.merge_cells("A6:H6")
    ws["A6"].fill = white_fill

    # ── KPI TILES helper ─────────────────────────────────────────────────────────
    # 3-row tile: accent bar (top) / label row / value row (bottom)
    # Reading order: accent colour → label → number  (natural top-to-bottom flow)
    # row_a : accent stripe
    # row_b : label (small text)
    # row_c : value (large bold number)

    def write_tile(row_a, row_b, row_c, col_start, col_end,
                   label_text, value, accent_hex, vfont):
        s = chr(ord("A") + col_start - 1)
        e = chr(ord("A") + col_end   - 1)

        # Accent bar — saturated stripe on top
        ws.merge_cells(f"{s}{row_a}:{e}{row_a}")
        ac = ws[f"{s}{row_a}"]
        ac.fill   = _fill(accent_hex)
        ac.border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=no_side)

        # Label row — off-white, small descriptive text  ← NOW SECOND (reads first after accent)
        ws.merge_cells(f"{s}{row_b}:{e}{row_b}")
        lc = ws[f"{s}{row_b}"]
        lc.value     = label_text
        lc.fill      = tile_spacer_fill
        lc.font      = tile_label_font
        lc.alignment = center
        lc.border    = Border(left=medium_side, right=medium_side, top=no_side, bottom=no_side)

        # Value row — white body, large bold number  ← NOW THIRD (punchline at bottom)
        ws.merge_cells(f"{s}{row_c}:{e}{row_c}")
        vc = ws[f"{s}{row_c}"]
        vc.value     = str(value)
        vc.fill      = white_fill
        vc.font      = vfont
        vc.alignment = center
        vc.border    = Border(left=medium_side, right=medium_side, top=no_side, bottom=medium_side)

    # Top row: neutral KPIs — dark grey numbers (calm, informational)
    neutral_font = Font(color=DARK_GREY, bold=True, size=16, name="Calibri")
    write_tile(7, 8, 9, 1, 2, "Total Terminating Services",  total_terminating_services,     MAROON_MID, neutral_font)
    write_tile(7, 8, 9, 3, 4, "Travelling to Yard",           services_direct_to_yard,        MAROON_MID, neutral_font)
    write_tile(7, 8, 9, 5, 6, "Shortest Dwell Prior to Yard", shortest_dwell_to_yard_display, MAROON_MID, neutral_font)
    write_tile(7, 8, 9, 7, 8, "Longest Dwell Prior to Yard",  longest_dwell_to_yard_display,  MAROON_MID, neutral_font)

    # Bottom row: risk KPIs — number colour matches the saturated accent so the rating is unmistakable
    high_font    = Font(color=HIGH_ACCENT,    bold=True, size=16, name="Calibri")
    medium_font  = Font(color=MEDIUM_ACCENT,  bold=True, size=16, name="Calibri")
    low_font     = Font(color=LOW_ACCENT,     bold=True, size=16, name="Calibri")
    minimum_font = Font(color=MINIMUM_ACCENT, bold=True, size=16, name="Calibri")

    write_tile(11, 12, 13, 1, 2, "High Risk Services",    high_risk_services,    HIGH_ACCENT,    high_font)
    write_tile(11, 12, 13, 3, 4, "Medium Risk Services",  medium_risk_services,  MEDIUM_ACCENT,  medium_font)
    write_tile(11, 12, 13, 5, 6, "Low Risk Services",     low_risk_services,     LOW_ACCENT,     low_font)
    write_tile(11, 12, 13, 7, 8, "Minimum Risk Services", minimum_risk_services, MINIMUM_ACCENT, minimum_font)

    # ── ROW 15+: DQ Observations (A–E) + Risk Legend (F–H) ──────────────────────
    # DQ gets 5 cols so warnings fit on one line; legend gets 3 cols (label + description merged)
    DQ_START = 15
    ws.row_dimensions[DQ_START].height = 13

    ws.merge_cells(f"A{DQ_START}:E{DQ_START}")
    c = ws[f"A{DQ_START}"]
    c.value     = "Data Quality Observations"
    c.fill      = section_fill
    c.font      = section_font
    c.alignment = left

    ws.merge_cells(f"F{DQ_START}:H{DQ_START}")
    c = ws[f"F{DQ_START}"]
    c.value     = "Risk Legend"
    c.fill      = section_fill
    c.font      = section_font
    c.alignment = left

    # Legend entries: label in F, description in G–H
    legend_entries = [
        (HIGH_C,    HIGH_ACCENT,    "High Risk",    "≤ 30s dwell to yard"),
        (MEDIUM_C,  MEDIUM_ACCENT,  "Medium Risk",  "30s < dwell ≤ 90s to yard"),
        (LOW_C,     LOW_ACCENT,     "Low Risk",     "90s < dwell ≤ 150s to yard"),
        (MINIMUM_C, MINIMUM_ACCENT, "Minimum Risk", "Not to yard / > 150s"),
    ]

    # Build warning messages — no wrap, so keep them concise as they already are
    warning_msgs = []
    if not one_sec_summary.empty:
        for _, r in one_sec_summary.iterrows():
            platform_text = ", ".join(r["Platforms"]) if r["Platforms"] else "unknown platform"
            warning_msgs.append((
                f"⚠  {r['Terminating Station']} — {int(r['Count'])} service(s) end at "
                f"{platform_text} with a 1-second dwell. Please review operational suitability.",
                False
            ))
    else:
        warning_msgs.append(("✔  No 1-second dwell anomalies detected.", True))

    content_rows = max(len(warning_msgs), len(legend_entries))

    for i in range(content_rows):
        row = DQ_START + 1 + i
        ws.row_dimensions[row].height = 14

        # DQ column A–E: no wrap, single line
        ws.merge_cells(f"A{row}:E{row}")
        wc = ws[f"A{row}"]
        if i < len(warning_msgs):
            msg_text, is_ok = warning_msgs[i]
            wc.value  = msg_text
            wc.font   = ok_font if is_ok else warning_font
            wc.fill   = warning_bg_fill
            wc.border = warn_border
        else:
            wc.fill   = warning_bg_fill
            wc.border = thin_box
        wc.alignment = left_nowrap   # NO wrap — single line per warning

        # Legend F (label) + G–H (description)
        if i < len(legend_entries):
            body_hex, accent_hex, lbl, desc = legend_entries[i]
            leg_fill    = _fill(body_hex)
            accent_fill = _fill(accent_hex)

            ws["F" + str(row)].fill   = accent_fill
            ws["F" + str(row)].value  = lbl
            ws["F" + str(row)].font   = Font(color=WHITE, bold=True, size=8, name="Calibri")
            ws["F" + str(row)].alignment = left_nowrap
            ws["F" + str(row)].border = thin_box

            ws.merge_cells(f"G{row}:H{row}")
            dc = ws[f"G{row}"]
            dc.value     = desc
            dc.fill      = leg_fill
            dc.font      = legend_font
            dc.alignment = left_nowrap
            dc.border    = thin_box

    # ── Note row ─────────────────────────────────────────────────────────────────
    note_row = DQ_START + 1 + content_rows
    ws.row_dimensions[note_row].height = 9
    ws.merge_cells(f"A{note_row}:H{note_row}")
    nc = ws[f"A{note_row}"]
    nc.value     = "Full terminating station details available in the output sheet"
    nc.font      = note_font
    nc.alignment = center

    # ── Top-N table ──────────────────────────────────────────────────────────────
    table_title_row = note_row + 1
    ws.row_dimensions[table_title_row].height = 13

    ws.merge_cells(f"A{table_title_row}:H{table_title_row}")
    c = ws[f"A{table_title_row}"]
    c.value     = f"Top {TOP_N} Terminating Stations"
    c.fill      = section_fill
    c.font      = section_font
    c.alignment = left

    headers = [
        "Station", "Terminating\nTrains", "Trains to\nYard",
        "Min. Dwell\nTo Yard (s)", "High Risk", "Medium Risk", "Low Risk", "Minimum Risk"
    ]

    hdr_row = table_title_row + 1
    ws.row_dimensions[hdr_row].height = 24

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.fill      = table_hdr_fill
        cell.font      = table_header_font
        cell.alignment = center
        cell.border    = thin_box

    row_ptr = hdr_row + 1
    if not station_summary.empty:
        for idx, (_, r) in enumerate(station_summary.iterrows()):
            ws.row_dimensions[row_ptr].height = 26

            is_alt    = (idx % 2 == 1)
            base_fill = alt_row_fill if is_alt else white_fill

            c1 = ws.cell(row=row_ptr, column=1, value=r["Terminating Station"])
            c1.font      = station_font
            c1.alignment = left
            c1.border    = thin_box
            c1.fill      = base_fill

            for col_idx, val in [
                (2, int(r["Terminating_Trains"])),
                (3, int(r["Trains_to_Yard"])),
                (4, int(r["Min_Dwell_To_Yard_Sec"]))
            ]:
                c = ws.cell(row=row_ptr, column=col_idx, value=val)
                c.alignment = center
                c.font      = data_font
                c.border    = thin_box
                c.fill      = base_fill

            risk_cols = [
                (5, r["High Risk"],    HIGH_RISK_FILL),
                (6, r["Medium Risk"],  MEDIUM_RISK_FILL),
                (7, r["Low Risk"],     LOW_RISK_FILL),
                (8, r["Minimum Risk"], MINIMUM_RISK_FILL),
            ]
            for col_idx, val, rfill in risk_cols:
                c = ws.cell(row=row_ptr, column=col_idx, value=val)
                c.fill      = rfill
                c.font      = Font(size=8, name="Calibri")
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.border    = thin_box

            row_ptr += 1
    else:
        ws.merge_cells(f"A{row_ptr}:H{row_ptr}")
        c = ws[f"A{row_ptr}"]
        c.value     = "No terminating station data available"
        c.alignment = center
        c.border    = thin_box

    last_used_row = max(row_ptr, hdr_row + 2)
    ws.print_area = f"A1:H{last_used_row}"


# ============================================================
# Output Sheet
# ============================================================
def create_output_sheet_inline(wb, df, sheet_name):
    ws = wb.create_sheet(title=sheet_name)

    df = df.copy()

    def map_day_display(day_raw):
        if day_raw in ["M______", "_T_____", "__W____", "___T___"]:
            return "Mon-Thu"
        elif day_raw == "____F__":
            return "Fri"
        elif day_raw == "_____S_":
            return "Sat"
        elif day_raw == "______S":
            return "Sun"
        return ""

    df["Day"] = df["Day"].apply(map_day_display)

    header_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
    sep_fill = PatternFill(start_color=SEPARATOR_GREY_HEX, end_color=SEPARATOR_GREY_HEX, fill_type="solid")

    thick = Side(style="medium")
    header_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    maroon_fill = PatternFill(start_color="F4E6E6", end_color="F4E6E6", fill_type="solid")

    centre_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)
    hidden_font = Font(color="F4E6E6")

    visible_cols = [
        c for c in df.columns
        if c not in {"Platform", "__pair_id", "SummaryPrevStation", "SummaryPrevDwellSec"}
    ]

    for col_idx, col in enumerate(visible_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centre_align
        cell.border = header_border

    ws.row_dimensions[1].height = HEADER_HEIGHT

    dv_yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_yes_no)

    col_index = {name: (i + 1) for i, name in enumerate(visible_cols)}
    highlight_cols = ["Terminating Station", "Dwell", "To Yard?", "Additional Dwell (en route)"]
    toyard_col = col_index.get("To Yard?")

    pair_rows = {}

    for excel_row, (_, row) in enumerate(df.iterrows(), start=2):
        is_blank = all(
            pd.isna(row[c]) or str(row[c]) == ""
            for c in visible_cols
        )

        for col_name in visible_cols:
            c_idx = col_index[col_name]
            val = row[col_name]

            cell = ws.cell(row=excel_row, column=c_idx)
            cell.alignment = centre_align

            if col_name in highlight_cols:
                cell.fill = maroon_fill

            if is_blank:
                cell.fill = sep_fill
                cell.value = None
                continue

            if col_name == "Dwell":
                cell.number_format = DWELL_NUMBER_FORMAT
                if pd.notna(val) and str(val) != "":
                    cell.value = excel_duration(val)
                else:
                    cell.value = None
                continue

            cell.value = None if (pd.isna(val) if isinstance(val, float) else False) else val

        if not is_blank and toyard_col is not None:
            ws.cell(row=excel_row, column=toyard_col).number_format = "@"
            dv_yes_no.add(ws.cell(row=excel_row, column=toyard_col))

        pid = row.get("__pair_id", None)
        if pd.notna(pid):
            pid = int(pid)
            pair_rows.setdefault(pid, []).append(excel_row)

    cols_to_hide = ["Terminating Station", "Dwell", "To Yard?", "Additional Dwell (en route)"]

    for pid, rows in pair_rows.items():
        if len(rows) != 2:
            continue
        r_top, r_bot = rows[0], rows[1]
        for col_name in cols_to_hide:
            if col_name not in col_index:
                continue
            c_idx = col_index[col_name]
            ws.cell(row=r_bot, column=c_idx).font = hidden_font

    thin = Side(style="thin")

    def apply_box_border(row_start, row_end):
        for r in range(row_start, row_end + 1):
            for c in range(1, len(visible_cols) + 1):
                cell = ws.cell(row=r, column=c)
                left = thin if c == 1 else None
                right = thin if c == len(visible_cols) else None
                top = thin if r == row_start else None
                bottom = thin if r == row_end else None
                cell.border = Border(
                    left=left or cell.border.left,
                    right=right or cell.border.right,
                    top=top or cell.border.top,
                    bottom=bottom or cell.border.bottom
                )

    for pid, rows in pair_rows.items():
        if len(rows) == 2:
            apply_box_border(rows[0], rows[1])

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        pid = row.get("__pair_id", None)
        if pd.isna(pid):
            is_blank = all(
                pd.isna(row[c]) or str(row[c]) == ""
                for c in visible_cols
            )
            if not is_blank:
                apply_box_border(r_idx, r_idx)

    widths = {
        "Run": 6, "TID": 8, "Day": 10, "Start": 8,
        "Org": ORG_DEST_WIDTH, "Dest": ORG_DEST_WIDTH, "Finish": 8,
        "Terminating Station": 24, "Dwell": 12,
        "To Yard?": 12, "Additional Dwell (en route)": 28
    }

    for col_name, width in widths.items():
        if col_name in col_index:
            col_letter = ws.cell(row=1, column=col_index[col_name]).column_letter
            ws.column_dimensions[col_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


# ============================================================
# Excel writer
# ============================================================
def write_excel(df_nursery, out_path, timetable_input, qa_info):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    day_order = ["Mon-Thu", "Fri", "Sat", "Sun"]
    available_day_groups = []

    if df_nursery["Day"].isin(["M______", "_T_____", "__W____", "___T___"]).any():
        available_day_groups.append("Mon-Thu")
    if (df_nursery["Day"] == "____F__").any():
        available_day_groups.append("Fri")
    if (df_nursery["Day"] == "_____S_").any():
        available_day_groups.append("Sat")
    if (df_nursery["Day"] == "______S").any():
        available_day_groups.append("Sun")

    available_day_groups = [d for d in day_order if d in available_day_groups]

    summary_sheet_names = []
    daily_outputs = {}

    for day_group in available_day_groups:
        daily_outputs[day_group] = build_final_output(df_nursery, day_filter=day_group)

    for day_group in available_day_groups:
        create_output_sheet_inline(wb, daily_outputs[day_group], sheet_name=f"Output ({day_group})")

    for day_group in available_day_groups:
        create_summary_sheet(wb, daily_outputs[day_group], sheet_name=f"Summary ({day_group})")
        summary_sheet_names.append(f"Summary ({day_group})")

    try:
        wb.save(out_path)
        print("Saved:", out_path)

        excel = None
        wb_excel = None

        try:
            import win32com.client
            from PyPDF2 import PdfMerger

            time.sleep(1)

            out_path_clean = os.path.normpath(os.path.abspath(out_path))

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb_excel = excel.Workbooks.Open(out_path_clean)

            pdf_stem = os.path.splitext(os.path.basename(out_path_clean))[0].replace("TerminatingTrainOutput-","")
            final_pdf_path = os.path.normpath(os.path.join(os.path.dirname(out_path_clean), f"TerminatingTrainOutputSummary-{pdf_stem}.pdf"))
            
            temp_pdf_files = []

            for sheet_name in summary_sheet_names:
                try:
                    sheet = wb_excel.Sheets(sheet_name)
                    sheet.Activate()

                    generated_text = datetime.now().strftime("%d/%m/%Y %H:%M")
                    excluded_count = qa_info["excluded_count"]
                    excluded_prefixes = qa_info["excluded_prefixes"]

                    if excluded_count > 0:
                        prefix_list = ", ".join(sorted(excluded_prefixes))
                        left_footer_text = (
                            f"&8&K666666QA: {excluded_count} service(s) excluded ({prefix_list})\n"
                            f"Generated: {generated_text}\n"
                            f"Timetable: {timetable_input}"
                        )
                    else:
                        left_footer_text = (
                            f"&8&K666666QA: Data integrity check passed\n"
                            f"Generated: {generated_text}\n"
                            f"Timetable: {timetable_input}"
                        )

                    right_footer_text = (
                        f'&8&K666666&"Arial,Bold"Master Train Planning\n'
                        f'&9&KD32F2F&"Arial,Bold"Queensland Rail'
                    )

                    sheet.PageSetup.Orientation = 2
                    sheet.PageSetup.Zoom = False
                    sheet.PageSetup.FitToPagesWide = 1
                    sheet.PageSetup.FitToPagesTall = 1
                    sheet.PageSetup.CenterHorizontally = True
                    sheet.PageSetup.LeftMargin = excel.InchesToPoints(0.3)
                    sheet.PageSetup.RightMargin = excel.InchesToPoints(0.3)
                    sheet.PageSetup.TopMargin = excel.InchesToPoints(0.35)
                    sheet.PageSetup.BottomMargin = excel.InchesToPoints(0.35)
                    sheet.PageSetup.FooterMargin = excel.InchesToPoints(0.2)
                    sheet.PageSetup.LeftFooter = left_footer_text
                    sheet.PageSetup.RightFooter = right_footer_text
                    sheet.PageSetup.CenterFooter = ""

                    safe_sheet_name = re.sub(r'[\\/:*?"<>|]', "_", sheet_name)
                    temp_pdf = os.path.normpath(out_path_clean.replace(".xlsx", f"_{safe_sheet_name}.pdf"))

                    sheet.ExportAsFixedFormat(Type=0, Filename=temp_pdf)
                    temp_pdf_files.append(temp_pdf)

                except Exception as e:
                    print(f"Failed exporting {sheet_name}:", e)

            if temp_pdf_files:
                merger = PdfMerger()
                for pdf in temp_pdf_files:
                    merger.append(pdf)
                merger.write(final_pdf_path)
                merger.close()
                for pdf in temp_pdf_files:
                    try:
                        os.remove(pdf)
                    except Exception:
                        pass
                print("PDF exported:", final_pdf_path)

        except Exception as e:
            print("PDF export failed:", e)

        finally:
            try:
                if wb_excel:
                    wb_excel.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel:
                    excel.Quit()
            except Exception:
                pass

        if AUTO_OPEN_AFTER_SAVE:
            os.startfile(out_path)

    except PermissionError:
        msg = (
            "The output file is currently open.\n\n"
            "Please close the Excel file and run the script again.\n\n"
            f"File:\n{out_path}"
        )
        print(msg)
        show_info_scroll_safe("File Open Error", msg)
        return


# ============================================================
# MAIN
# ============================================================
def run_terminating_train_output(path, mypath=None):
    try:
        df_nursery, qa_info = build_nursery_rows_from_rsx(path)

        

       
        input_path = Path(path)
        out_path = str(input_path.with_name(OUTPUT_NAME_SUFFIX + input_path.stem + ".xlsx"))

        write_excel(df_nursery, out_path, input_path.stem, qa_info)

    except Exception as e:
        logging.error(traceback.format_exc())
        #show_info_scroll_safe("Script Error", f"Error:\n\n{str(e)}")
        return


if __name__ == "__main__":
    app = QApplication.instance() or QApplication(sys.argv)
    path = select_file(
        caption='Select RSX file',
        directory='',
        filter_str='RSX Files (*.rsx);;All Files (*.*)'
    )
    if path:
        run_terminating_train_output(path)
