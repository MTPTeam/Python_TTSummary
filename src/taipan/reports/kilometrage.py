"""
kilometrage.py
Python port of the VBA ImportDataAndCreateOutput macro.
Reads a RailSys kilometrage Excel export, builds a formatted summary workbook, and saves it as "Kilometrage Output.xlsx" in the same directory as the input file.
"""


import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import QApplication
from taipan.gui.base import select_file, show_info


COLUMN_MAPPINGS = {
    "Number":                    1,
    "Train formation":           2,
    "Line":                      3,
    "Length of train run [km]":  4,
    "DoO":                       5,
    "Scheduled travel time":     6,
}



def rgb(r, g, b) -> str:
    """Return openpyxl ARGB hex string."""
    return f"FF{r:02X}{g:02X}{b:02X}"

THIN  = Side(style="thin")
THICK = Side(style="double")

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def double_border():
    return Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

def apply_fill(ws, cell_range, r, g, b):
    fill = PatternFill("solid", fgColor=rgb(r, g, b))
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill

def apply_font(ws, cell_range, bold=False, size=11, color="FF000000"):
    font = Font(bold=bold, size=size, color=color)
    for row in ws[cell_range]:
        for cell in row:
            cell.font = font

def apply_alignment(ws, cell_range, horizontal="general"):
    align = Alignment(horizontal=horizontal)
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = align

def set_thin_borders(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border()

def set_outer_double_border(ws, cell_range):
    """Apply a double border around the outside of a range."""
    rows = list(ws[cell_range])
    top_row, bot_row = rows[0], rows[-1]
    for cell in top_row:
        b = cell.border.copy()
        cell.border = Border(left=b.left, right=b.right, top=THICK, bottom=b.bottom)
    for cell in bot_row:
        b = cell.border.copy()
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THICK)
    for row in rows:
        lc, rc = row[0], row[-1]
        bl = lc.border.copy()
        lc.border = Border(left=THICK, right=bl.right, top=bl.top, bottom=bl.bottom)
        br = rc.border.copy()
        rc.border = Border(left=br.left, right=THICK, top=br.top, bottom=br.bottom)

def white_bottom_border(ws, cell_range):
    white = Side(style="thin", color="FFFFFFFF")
    for row in ws[cell_range]:
        for cell in row:
            b = cell.border.copy()
            cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=white)


def autofit_selected_columns(ws, columns):
    for col_letter in columns:
        max_length = 0

        for row in ws.iter_rows():
            cell = row[ord(col_letter) - ord('A')]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2


# section layout
SECTIONS = [
    ("TOTAL KM's",   1),
    ("NGR KM's",     8),
    ("EMU/SMU KM's", 18),
    ("IMU KM's",     28),
    ("Tot. QR KM's", 38),
]

DAY_HEADERS = {
    3: "Mon - Thurs",
    4: "Fri",
    5: "Sat",
    6: "Sun",
    7: "Total Weekly Km's",
}

# SUMIFS day codes used in formulas
DAY_CODES = {
    3: "Mo-Thu",
    4: "Fr",
    5: "Sa",
    6: "Sun",
}

# Train type filter for each section (None = no filter = Total)
SECTION_FILTERS = [None, "NGR", "EMU", "IMU", None]

# Col letters for formula building
COL = {3: "C", 4: "D", 5: "E", 6: "F", 7: "G"}

def empty_formula(col_letter, day_code, train_type=None):
    if train_type:
        return f'=SUMIFS(M:M,P:P,"Empty",N:N,"{day_code}",Q:Q,"{train_type}")'
    return f'=SUMIFS(M:M,P:P,"Empty",N:N,"{day_code}")'

def revenue_formula(col_letter, day_code, train_type=None):
    if train_type:
        return f'=SUMIFS(M:M,P:P,"Revenue",N:N,"{day_code}",Q:Q,"{train_type}")'
    return f'=SUMIFS(M:M,P:P,"Revenue",N:N,"{day_code}")'


def build_output(ws, last_row):
    for r in range(2, last_row + 1):
        ws.cell(r, 16).value = f'=IF(LEFT(K{r},1)="E","Empty","Revenue")'
        ws.cell(r, 17).value = f'=IF(P{r}="Revenue",MID(K{r},3,3),IF(P{r}="Empty",MID(K{r},9,3)))'

    acc_fmt  = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    int_fmt  = '0'
    for row in ws["A1:H43"]:
        for cell in row:
            cell.number_format = acc_fmt
    for addr in ["F15", "F25", "F35"]:
        ws[addr].number_format = int_fmt

    major_rows = [1, 8, 18, 28, 38]
    major_labels = ["TOTAL KM's", "NGR KM's", "EMU/SMU KM's", "IMU KM's", "Tot. QR KM's"]
    for row, label in zip(major_rows, major_labels):
        cell = ws.cell(row, 2)
        cell.value = label
        cell.font  = Font(bold=True, size=18, color="FFFFFFFF")

    ws.cell(1, 16).value = "Service Type"
    ws.cell(1, 17).value = "Train Type"
    apply_font(ws, "J1:Q1", bold=True)

    apply_fill(ws, "A8:H16",  185, 50,  50)
    apply_fill(ws, "A18:H26", 185, 50,  50)
    apply_fill(ws, "A28:H36", 185, 50,  50)
    apply_fill(ws, "A1:H6",   100, 100, 100)
    apply_fill(ws, "A38:H43", 100, 100, 100)
    apply_fill(ws, "B2:G5",   255, 255, 255)
    apply_fill(ws, "B9:G15",  255, 255, 255)
    apply_fill(ws, "B19:G25", 255, 255, 255)
    apply_fill(ws, "B29:G35", 255, 255, 255)
    apply_fill(ws, "B39:G42", 255, 255, 255)

    inner_ranges   = ["B2:G5",   "B9:G12",  "B19:G22", "B29:G32", "B39:G42"]
    divider_ranges = ["B3:G3",   "B10:G10", "B20:G20", "B30:G30", "B40:G40"]
    for ir, dr in zip(inner_ranges, divider_ranges):
        set_thin_borders(ws, ir)
        set_outer_double_border(ws, ir)
        white_bottom_border(ws, dr)

    minor_rows = [
        (5,  "TOTAL KM's"),
        (12, "TOTAL KM's"),
        (22, "TOTAL KM's"),
        (32, "TOTAL KM's"),
        (42, "TOTAL KM's"),
    ]
    for row, label in minor_rows:
        c = ws.cell(row, 2)
        c.value = label
        c.font  = Font(bold=True, size=16)

    apply_font(ws, "B5:G5",   bold=True, size=16)
    apply_font(ws, "B12:G12", bold=True, size=16)
    apply_font(ws, "B22:G22", bold=True, size=16)
    apply_font(ws, "B32:G32", bold=True, size=16)
    apply_font(ws, "B42:G42", bold=True, size=16)

    empty_label   = "KM's - Empty Services"
    revenue_label = "KM's - Revenue Services"

    for base in [1, 8, 18, 28, 38]:
        off = 2 if base == 1 else (1 if base == 8 else (2 if base == 18 else (2 if base == 28 else 2)))
        # rows: empty = base+2, revenue = base+3  
        e_row = {1: 3, 8: 10, 18: 20, 28: 30, 38: 40}[base]
        r_row = {1: 4, 8: 11, 18: 21, 28: 31, 38: 41}[base]
        ws.cell(e_row, 2).value = empty_label
        ws.cell(r_row, 2).value = revenue_label

    for base in [1, 8, 18, 28, 38]:
        header_row = {1: 2, 8: 9, 18: 19, 28: 29, 38: 39}[base]
        for col, label in DAY_HEADERS.items():
            c = ws.cell(header_row, col)
            c.value = label
            c.font  = Font(bold=True, size=9, color="FF000064")
            c.alignment = Alignment(horizontal="center")

    section_defs = [
        # (empty_row, revenue_row, total_row, train_type, total_col_empty_formula, total_col_rev_formula, total_col_total_formula)
        (3,  4,  5,  None,  "=G20+G30", "=G21+G31", "=G40+G41"),   # Total
        (10, 11, 12, "NGR", None, None, None),
        (20, 21, 22, "EMU", None, None, None),
        (30, 31, 32, "IMU", None, None, None),
        (40, 41, 42, None,  "=G20+G30", "=G21+G31", "=G40+G41"),   # Tot QR
    ]

    # Per-day formulas (cols C-F)
    for idx, (e_row, r_row, t_row, ttype, g_e, g_r, g_t) in enumerate(section_defs):
        for col in [3, 4, 5, 6]:
            day_code   = DAY_CODES[col]
            col_letter = COL[col]
            # Special total-column overrides for "Total" and "Tot. QR" sections
            is_total_section = ttype is None and idx in (0, 4)
            if is_total_section and col in [3, 4, 5, 6]:
                if idx == 4:  # Tot. QR = EMU + IMU
                    base_e = col_letter
                    ws.cell(e_row, col).value = f"={base_e}20+{base_e}30"
                    ws.cell(r_row, col).value = f"={base_e}21+{base_e}31"
                    ws.cell(t_row, col).value = f"={base_e}{e_row}+{base_e}{r_row}"
                    continue
            ws.cell(e_row, col).value = empty_formula(col_letter, day_code, ttype)
            ws.cell(r_row, col).value = revenue_formula(col_letter, day_code, ttype)
            ws.cell(t_row, col).value = f"={col_letter}{e_row}+{col_letter}{r_row}"

        # Weekly total column (G)
        if ttype is None and idx == 4:  # Tot. QR
            ws.cell(e_row, 7).value = "=G20+G30"
            ws.cell(r_row, 7).value = "=G21+G31"
        elif ttype is None:             # Grand Total
            ws.cell(e_row, 7).value = "=C3*4+D3+E3+F3"
            ws.cell(r_row, 7).value = "=C4*4+D4+E4+F4"
        else:
            ws.cell(e_row, 7).value = f"=C{e_row}*4+D{e_row}+E{e_row}+F{e_row}"
            ws.cell(r_row, 7).value = f"=C{r_row}*4+D{r_row}+E{r_row}+F{r_row}"
        ws.cell(t_row, 7).value = f"=G{e_row}+G{r_row}"

    extra_ranges  = ["F14:G15", "F24:G25", "F34:G35"]
    header_ranges = ["F14:G14", "F24:G24", "F34:G34"]
    yellow_cells  = ["F15", "F25", "F35"]
    fleet_labels  = ["# NGR Fleet", "# EMU/SMU Fleet", "# IMU Fleet"]
    yearly_rows   = [15, 25, 35]
    fleet_rows    = [14, 24, 34]
    g_rows        = [12, 22, 32]
    fleet_counts  = [72, 39, 21]

    for er, hr, yc, fl, yr, fr, gr, fc in zip(
        extra_ranges, header_ranges, yellow_cells,
        fleet_labels, yearly_rows, fleet_rows, g_rows, fleet_counts
    ):
        set_thin_borders(ws, er)
        apply_font(ws, f"G{yr}:G{yr}", bold=True, size=16)
        apply_font(ws, hr, bold=True, size=9, color="FF000064")
        ws.cell(yr, 7).value  = f"=G{gr}*52/F{yr}"
        ws.cell(fr, 7).value  = "Total Yearly Km / Unit"
        ws.cell(fr, 6).value  = fl
        ws.cell(yr, 6).value  = fc
        yc_cell = ws[yc]
        yc_cell.fill   = PatternFill("solid", fgColor=rgb(255, 255, 0))
        yc_cell.border = double_border()
        yc_cell.number_format = "0"
        apply_font(ws, f"F{yr}:G{yr}", bold=True, size=16)

    for col in range(3, 8):                              # C:G
        ws.column_dimensions[get_column_letter(col)].width = 16.5
    apply_alignment(ws, "C1:G43", horizontal="center")
    apply_alignment(ws, "J1:Q43", horizontal="center")
    ws.column_dimensions["A"].width = 3


    autofit_selected_columns(ws, ["B", "P"])


def main(path: str) -> None:
    base, ext   = os.path.splitext(path)
    out_path    = os.path.join(os.path.dirname(path), "Kilometrage Output.xlsx")

    df = pd.read_excel(path, sheet_name=0, header=0)
    last_row = len(df) + 1   # +1 for header row

    wb = Workbook()
    ws = wb.active

    # Copy mapped columns to output sheet at offset +9
    for col_name, out_col in COLUMN_MAPPINGS.items():
        if col_name in df.columns:
            dest_col = out_col + 9
            ws.cell(1, dest_col).value = col_name   # header
            for row_idx, value in enumerate(df[col_name], start=2):
                ws.cell(row_idx, dest_col).value = value

    # Build the summary section and formatting
    build_output(ws, last_row)
    wb.save(out_path)
    os.startfile(out_path)
    show_info("Success", "Kilometrage Calculated Successfully!")

if __name__ == "__main__":
    app = QApplication.instance() or QApplication(sys.argv)
    path = select_file(caption="Select Excel file", directory="", filter_str="Excel Files (*.xlsx *.xls);;All Files (*.*)")
    if path:
        main(path)
